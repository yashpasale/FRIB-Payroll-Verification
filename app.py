import tkinter as tk
from tkinter import filedialog
import os
import shutil
import pandas as pd
import re
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


app = tk.Tk()
app.title("FRIB Payroll Validation")
app.geometry("500x300")

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


def on_closing():
    clear_uploads_folder()
    app.destroy()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls', 'txt'}


def get_total_rows(filepath):
    try:
        df = pd.read_excel(filepath)
        return len(df)
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


def txt_to_excel(file_path):
    try:
        if os.path.exists(file_path):
            with open(file_path, 'r') as text_file:
                text_content = text_file.read()

            data = [re.split(r'\s+', line.strip()) for line in text_content.split('\n')]
            df = pd.DataFrame(data)
            excel_save_path = filedialog.asksaveasfilename(defaultextension='.xlsx',
                                                           filetypes=[("Excel Files", "*.xlsx")],
                                                           title="Save Excel File As")
            if excel_save_path:
                df.to_excel(excel_save_path, index=False, header=False)
                status_label.config(text="File converted to Excel successfully")
                clear_uploads_folder()
            else:
                status_label.config(text="Excel save cancelled")
        else:
            status_label.config(text="File not found")
    except Exception as e:
        print(f"An error occurred during conversion: {e}")
        status_label.config(text="Failed to convert file")


def upload_file():
    file_path = filedialog.askopenfilename()
    if file_path and os.path.exists(file_path):
        filename = os.path.basename(file_path)
        if allowed_file(filename):
            destination_path = os.path.join(UPLOAD_FOLDER, filename)
            shutil.copy(file_path, destination_path)
            upload_button.pack_forget()
            app.file_path = destination_path
            status_label.config(text="File uploaded successfully")
            continue_button.pack()
            row_count_button.pack(pady=5)
        else:
            status_label.config(text="Invalid file format")
    else:
        status_label.config(text="File not found")


def display_row_count():
    if app.file_path and os.path.exists(app.file_path):
        try:
            total_rows = get_total_rows(app.file_path)
            if total_rows is not None:
                status_label.config(text=f"File has {total_rows} rows")
            else:
                status_label.config(text="Error: Unable to get row count")
        except Exception as e:
            print(f"An error occurred while getting row count: {e}")
            status_label.config(text="Error: Unable to get row count")
    else:
        status_label.config(text="File path is not valid")


def upload_text_file():
    file_path = filedialog.askopenfilename()
    if file_path and os.path.exists(file_path):
        filename = os.path.basename(file_path)
        if allowed_file(filename):
            destination_path = os.path.join(UPLOAD_FOLDER, filename)
            shutil.copy(file_path, destination_path)
            status_label.config(text="Txt file Uploaded Successfully")
            app.file_path = destination_path
            txt_to_excel(destination_path)
        else:
            status_label.config(text="Invalid file format")
    else:
        status_label.config(text="File not found")


def clear_uploads_folder():
    for filename in os.listdir(UPLOAD_FOLDER):
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f"Error deleting file: {e}")


def run_process():
    if app.file_path and os.path.exists(app.file_path):
        try:
            # Read Excel file into a DataFrame
            df = pd.read_excel(app.file_path, header=None)
            if df.astype(str).apply(lambda x: x.str.contains('test', case=False)).any().any():
                df = df.drop(df[df.astype(str).apply(lambda x: x.str.contains('test', case=False)).any(axis=1)].index, axis=0).reset_index(drop=True)

            # Assign column headers
            headers = ["Person ID", "PERNR", "Sub Account", "Project", "Date", "hours", "wo", "REGU", "Others"]
            df.columns = headers
            df['hours'] = pd.to_numeric(df['hours'], errors='coerce')

            # Find rows where "Project" contains "2022FRIB"
            mask = df['Project'].str.contains('2022FRIB', na=False)

            # Rearrange rows where "Project" contains "2022FRIB"
            for index, row in df[mask].iterrows():
                project_value = row['Project']
                other_values = row.drop('Project').values.tolist()
                rearranged_row = other_values + [project_value]
                df.loc[index, :] = rearranged_row

            mask = df['Project'].str.contains('COVID', na=False)

            # Rearrange rows where "Project" contains "COVID"
            for index, row in df[mask].iterrows():
                project_value = row['Project']
                other_values = row.drop('Project').values.tolist()
                rearranged_row = other_values + [project_value]
                df.loc[index, :] = rearranged_row

            mask = df['Sub Account'].str.startswith('2').fillna(False)
            for index, row in df[mask].iterrows():
                sub_account_index = df.columns.get_loc('Sub Account')
                sub_account_value = row['Sub Account']
                df.iloc[index, sub_account_index] = None
                df.iloc[index, sub_account_index + 3:] = df.iloc[index, sub_account_index + 1:-2].values
                df.iloc[index, sub_account_index + 2] = sub_account_value
                df.at[index, 'Project'] = None

            # FILTERS
            df_filtered_by_standby2420 = df[df['Date'].astype(str).str.endswith('2420')]
            df_filtered_by_BlankSubaccounts = df[df['Sub Account'].isnull() & df['Project'].isnull()]
            df_filtered_by_RC113931 = df[df['Project'].str.contains('RC113931', na=False)]
            df_filtered_by_GA016641 = df[df['Project'].str.contains('GA016641', na=False)]
            df_filtered_by_GA016641 = df_filtered_by_GA016641[df_filtered_by_GA016641['Sub Account'] == 'NO_SUB-ACCOUNT']

            # Review Hours
            

            save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[("Excel files", "*.xlsx")], title="Save Modified File As")

            if save_path:
                # Open the Excel file for further formatting
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Original_Data'

                # Write DataFrame to the first sheet
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

                # Create a new sheet for filtered data
                ws_filtered1 = wb.create_sheet(title='Standby_paycode_2420')
                for r in dataframe_to_rows(df_filtered_by_standby2420, index=False, header=True):
                    ws_filtered1.append(r)

                # Apply formatting to the new sheet
                for column in ws_filtered1.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)  # Get column letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws_filtered1.column_dimensions[column_letter].width = adjusted_width

                    if column_letter == 'E':  # Assuming 'Date' column is the fifth column
                        for cell in column[1:]:
                            cell.number_format = '0'

                ws_filtered2 = wb.create_sheet(title='Filtered_by_BlankSubaccounts')
                # Write DataFrame to the sheet
                for r in dataframe_to_rows(df_filtered_by_BlankSubaccounts, index=False, header=True):
                    ws_filtered2.append(r)

                # Create a new sheet for RC113931 filtered data
                ws_filtered3 = wb.create_sheet(title='Filtered_by_RC113931')
                # Write DataFrame to the sheet
                for r in dataframe_to_rows(df_filtered_by_RC113931, index=False, header=True):
                    ws_filtered3.append(r)

                ws_filtered4 = wb.create_sheet(title='GA016641 & NO_Subacounts')
                for r in dataframe_to_rows(df_filtered_by_GA016641, index=False, header=True):
                    ws_filtered4.append(r)

                # Apply formatting to all sheets
                for ws in wb.sheetnames:
                    ws_obj = wb[ws]
                    for column in ws_obj.columns:
                        max_length = 0
                        column_letter = openpyxl.utils.get_column_letter(column[0].column)  # Get column letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2
                        ws_obj.column_dimensions[column_letter].width = adjusted_width

                        if column_letter == 'E':  # Assuming 'Date' column is the fifth column
                            for cell in column[1:]:
                                cell.number_format = '0'
                # Save the workbook
                wb.save(save_path)

                status_label.config(text=f"File modified successfully and saved as {save_path}")
                clear_uploads_folder()
            else:
                status_label.config(text="Save cancelled")
        except Exception as e:
            print(f"An error occurred while processing the file: {e}")
            status_label.config(text="Error: Failed to process the file")
    else:
        status_label.config(text="File path is not valid")



def review_hours():
    if app.file_path and os.path.exists(app.file_path):
        try:
            df = pd.read_excel(app.file_path)

            # Apply formula to extract date from column E and put it in column K
            df['Date Extracted'] = df['Date'].apply(lambda x: str(x)[:8])

            # Create a pivot table
            pivot_table = df.pivot_table(index=['Person ID', 'Date Extracted'], values='hours', aggfunc='sum')

            # Filter employees with 16 hours on the same day
            employees_with_16_hours = pivot_table[pivot_table['hours'] == 16]

            status_label.config(text=f"Review completed. Employees with 16 hours on the same day: {employees_with_16_hours}")

        except Exception as e:
            print(f"An error occurred during review: {e}")
            status_label.config(text="Error: Review failed")


header_label = tk.Label(app, text="FRIB Payroll Validation", font=("Arial", 16, "bold"))
header_label.pack(pady=10)

upload_frame = tk.Frame(app)
upload_frame.pack(pady=10)

upload_label = tk.Label(upload_frame, text="Upload Excel File")
upload_label.grid(row=0, column=0, padx=10)

upload_button = tk.Button(upload_frame, text="Browse", command=upload_file)
upload_button.grid(row=0, column=1, padx=10)

status_label = tk.Label(app, text="")
status_label.pack(pady=5)

continue_button = tk.Button(app, text="Continue", command=run_process)
continue_button.pack_forget()

txt_to_excel_button = tk.Button(app, text="Convert Text to Excel", command=upload_text_file)
txt_to_excel_button.pack(pady=5)

row_count_button = tk.Button(app, text="Display Row Count", command=display_row_count)
row_count_button.pack_forget()



app.file_path = None
app.protocol("WM_DELETE_WINDOW", on_closing)

if __name__ == "__main__":
    app.mainloop()
